#06/03/2025
#@PLima


import random
import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
from collections import defaultdict

# Constantes
ARQUIVO_EXCEL_MEGA_SENA = "Mega-Sena.xlsx"
ARQUIVO_EXCEL_APOSTAS = "apostas.xlsx"
NUMERO_DE_NUMEROS_POR_APOSTA = 6
INTERVALO_NUMEROS = range(1, 61)
NUMERO_MAXIMO_TENTATIVAS_GERAR_APOSTA = 10
ATUALIZACAO_PLANILHA_INTERVALO = 5000
TOP_N = 10
BOTTOM_N = 10
MAX_TENTATIVAS_VALIDACAO = 5  # Máximo de tentativas para gerar uma aposta válida

def agora():
    agora = datetime.now()
    agora = agora.strftime("%Y-%m-%d %H-%M-%S")
    return str(agora)

def calcular_frequencia_numeros(arquivo_excel: str) -> dict[int, int]:
    """Calcula a frequência dos números."""
    try:
        wb = load_workbook(arquivo_excel)
        ws = wb.active
        frequencias = {i: 0 for i in range(1, 61)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            for numero in row[2:8]:
                if isinstance(numero, int) and 1 <= numero <= 60:
                    frequencias[numero] += 1
        return frequencias
    except FileNotFoundError:
        print(f"Arquivo não encontrado: {arquivo_excel}")
        return {}
    except Exception as e:
        print(f"Erro ao calcular a frequência dos números: {e}")
        return {}

def gerar_aposta_ponderada(frequencias):
    """Gera uma aposta com base na frequência ponderada dos números."""
    # Cria uma lista de números ponderada pela frequência
    numeros_ponderados = []
    for numero, frequencia in frequencias.items():
        numeros_ponderados.extend([numero] * frequencia)  # Adiciona o número 'frequencia' vezes

    # Se a lista ponderada estiver vazia (sem dados), retorna None
    if not numeros_ponderados:
        return None

    # Garante que haja pelo menos 6 números únicos
    if len(set(numeros_ponderados)) < 6:
        numeros_faltantes = random.sample([n for n in range(1, 61) if n not in set(numeros_ponderados)], 6 - len(set(numeros_ponderados)))
        numeros_ponderados.extend(numeros_faltantes)

    # Seleciona 6 números aleatórios da lista ponderada
    aposta = random.sample(numeros_ponderados, 6)
    return sorted(aposta)


def gerar_aposta_analisada():
    """Gera uma aposta com análise e probabilidade ponderada."""
    frequencias = calcular_frequencia_numeros(ARQUIVO_EXCEL_MEGA_SENA)
    if not frequencias:
        return None

    aposta = gerar_aposta_ponderada(frequencias)
    return aposta

def tem_sequencia_consecutiva(numeros: list[int]) -> bool:
    """Verifica se há 4 ou mais números consecutivos."""
    numeros.sort()
    contador = 1
    for i in range(len(numeros) - 1):
        if numeros[i] + 1 == numeros[i + 1]:
            contador += 1
            if contador >= 4:
                return True
        else:
            contador = 1
    return False

def muitos_multiplos_de_5(numeros: list[int]) -> bool:
    """Evita que mais da metade dos números sejam múltiplos de 5."""
    multiplos = [n for n in numeros if n % 5 == 0]
    return len(multiplos) >= 3

def padrao_visual_obvio(numeros: list[int]) -> bool:
    """Evita padrões como todos números da mesma dezena."""
    dezenas = {n // 10 for n in numeros}
    return len(dezenas) <= 2

def salvar_aposta_excel(numeros: list[int]) -> None:
    """Salva a aposta na planilha "apostas.xlsx"."""
    try:
        data_atual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if os.path.exists(ARQUIVO_EXCEL_APOSTAS):
            wb = load_workbook(ARQUIVO_EXCEL_APOSTAS)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Data", "N1", "N2", "N3", "N4", "N5", "N6"])
        ws.append([data_atual] + numeros)
        wb.save(ARQUIVO_EXCEL_APOSTAS)
    except Exception as e:
        print(f"Erro ao salvar a aposta no Excel: {e}")

def exibir_aposta() -> None:
    print(f'{agora()} - def exibir_aposta() - Inicio')
    """Gera, valida, exibe e salva a aposta."""
    tentativas = 0
    aposta = None

    while tentativas < MAX_TENTATIVAS_VALIDACAO:
        aposta_temp = gerar_aposta_analisada()
        if aposta_temp and not (tem_sequencia_consecutiva(aposta_temp) or muitos_multiplos_de_5(aposta_temp) or padrao_visual_obvio(aposta_temp)):
            aposta = aposta_temp
            print(f'{agora()} - aposta: {aposta}')
            break  # A aposta é válida, sai do loop
        tentativas += 1

    if aposta:
        label_resultado.config(text=f"Sua aposta: {aposta}")
        salvar_aposta_excel(aposta)
        atualizar_total_sorteios()
        atualizar_janela_planilha()
        print(f'{agora()} - def exibir_aposta() - Fim')
    else:
        label_resultado.config(text="Não foi possível gerar uma aposta válida após várias tentativas.")

def ao_fechar():
    """Confirma o fechamento do aplicativo."""
    resultado = messagebox.askyesno("Confirmação", "Tem certeza de que deseja fechar o aplicativo?")
    if resultado:
        print(f"{agora()} - Fechando aplicativo...")
        janela.destroy()

def contar_sorteios_excel() -> int:
    """Conta o número de sorteios na planilha."""
    try:
        if os.path.exists(ARQUIVO_EXCEL_APOSTAS):
            wb = load_workbook(ARQUIVO_EXCEL_APOSTAS)
            ws = wb.active
            return ws.max_row - 1
        else:
            return 0
    except Exception as e:
        print(f"Erro ao contar os sorteios no Excel: {e}")
        return 0

def atualizar_total_sorteios() -> None:
    """Atualiza o label com o total de sorteios."""
    total_sorteios = contar_sorteios_excel()
    label_total_sorteios.config(text=f"Total de sorteios: {total_sorteios}")

def exibir_janela_planilha() -> None:
    """Cria e exibe a janela da planilha."""
    global janela_planilha, tree
    janela_planilha = tk.Toplevel(janela)
    janela_planilha.title("Visualização da Planilha de Apostas")

    janela_planilha.geometry("600x400")

    tree = ttk.Treeview(janela_planilha, columns=("Data", "N1", "N2", "N3", "N4", "N5", "N6"), show="headings")

    tree.column("Data", width=150)
    tree.column("N1", width=50)
    tree.column("N2", width=50)
    tree.column("N3", width=50)
    tree.column("N4", width=50)
    tree.column("N5", width=50)
    tree.column("N6", width=50)

    tree.heading("Data", text="Data")
    tree.heading("N1", text="N1")
    tree.heading("N2", text="N2")
    tree.heading("N3", text="N3")
    tree.heading("N4", text="N4")
    tree.heading("N5", text="N5")
    tree.heading("N6", text="N6")
    tree.pack(expand=True, fill="both")

    atualizar_janela_planilha()

def atualizar_janela_planilha() -> None:
    """Atualiza os dados exibidos na janela da planilha."""
    try:
        if os.path.exists(ARQUIVO_EXCEL_APOSTAS):
            wb = load_workbook(ARQUIVO_EXCEL_APOSTAS)
            ws = wb.active
            dados = list(ws.values)

            for item in tree.get_children():
                tree.delete(item)

            for row in reversed(dados[1:]):
                tree.insert("", "end", values=row)

        janela_planilha.after(ATUALIZACAO_PLANILHA_INTERVALO, atualizar_janela_planilha)
    except Exception as e:
        print(f"Erro ao atualizar a janela da planilha: {e}")

if __name__ == "__main__":
    print(f'\n{agora()} - inicio.py - __main__ - Inicio')

    janela = tk.Tk()
    janela.title("Gerador de Apostas Mega-Sena")

    janela.geometry("500x250")
    janela.maxsize(500, 250)
    janela.protocol("WM_DELETE_WINDOW", ao_fechar)

    label_resultado = tk.Label(janela, text="Clique no botão para gerar uma aposta", font=("Arial", 12))
    label_resultado.place(relx=0.5, y=75, anchor='center')

    botao_gerar = tk.Button(janela, text="Gerar Aposta", command=exibir_aposta, font=("Arial", 12))
    botao_gerar.place(relx=0.5, y=175, anchor='center')

    label_total_sorteios = tk.Label(janela, text="", font=("Arial", 10))
    label_total_sorteios.place(relx=1.0, rely=1.0, anchor='se')

    exibir_janela_planilha()

    atualizar_total_sorteios()

    janela.mainloop()

    print(f'\n{agora()} - inicio.py - __main__ - Fim')